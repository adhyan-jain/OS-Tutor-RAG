"""Shared RAGAS scoring logic, used by both RAGPipeline.evaluate() (single-run,
auto-triggered after ingest) and eval/ragas_eval.py (mass comparison across
every technique combination). Single source of truth so both stay consistent.
"""

from __future__ import annotations

import json
import math
import os
from pathlib import Path
from typing import TYPE_CHECKING

# Force local-only HF Hub access: all models used here (bge-*) are already
# cached locally, and this environment's network to huggingface.co has been
# unreliable (has caused multi-minute hangs on cache-check requests).
os.environ.setdefault("HF_HUB_OFFLINE", "1")
os.environ.setdefault("TRANSFORMERS_OFFLINE", "1")

from src.config import PipelineConfig

if TYPE_CHECKING:
    from src.pipeline import RAGPipeline

METRIC_NAMES = [
    "context_precision",
    "context_recall",
    "faithfulness",
    "answer_relevancy",
    "answer_correctness",
]


def load_eval_set(eval_set_path: Path, num_questions: int | None = None) -> list[dict]:
    """Load eval examples from an eval_set.json file.

    Args:
        eval_set_path: Path to a JSON eval set (see eval/eval_set.json).
        num_questions: If given, only the first N examples are returned --
            lets you analyze on a smaller/larger slice without editing the
            eval set file itself.

    Returns:
        The list of example dicts (question, ground_truth_answer,
        source_chunk_ids, topic).
    """
    with open(eval_set_path) as f:
        data = json.load(f)
    examples = data["examples"]
    return examples[:num_questions] if num_questions is not None else examples


def score_pipeline(pipeline: "RAGPipeline", eval_set: list[dict]) -> tuple[dict[str, float], list[dict]]:
    """Score a pipeline's answers against a labeled eval set using RAGAS.

    Scores each row against each metric using the metric's synchronous
    ``.score(row)`` method, one call at a time. This deliberately bypasses
    ragas's own ``evaluate()``/executor machinery: that executor fires many
    concurrent async sub-calls per row (per ragas's internal metric
    implementations, not something RunConfig(max_workers=1) actually limits),
    which repeatedly deadlocked a single local Ollama server on an 8GB-VRAM
    GPU (confirmed via nvidia-smi/ollama ps/thread CPU-time sampling showing
    zero progress). Fully sequential scoring is slower but reliable here.

    Args:
        pipeline: The configured (and already-ingested) RAGPipeline instance
            to evaluate.
        eval_set: Eval examples as returned by load_eval_set().

    Returns:
        A tuple of (mean scores dict, per-question detail rows list). Mean
        scores are NaN-safe: a metric that fails to parse for some rows
        (local LLMs don't always follow ragas's expected structured output)
        is averaged over only the rows where it succeeded, not zeroed out
        or left to poison the whole mean.
    """
    import asyncio

    from langchain_community.chat_models import ChatOllama
    from langchain_community.embeddings import HuggingFaceEmbeddings
    from ragas.embeddings import LangchainEmbeddingsWrapper
    from ragas.llms import LangchainLLMWrapper
    from ragas.metrics import (
        answer_correctness,
        answer_relevancy,
        context_precision,
        context_recall,
        faithfulness,
    )
    from ragas.run_config import RunConfig

    # metric.score() calls asyncio.get_event_loop() internally, which Python
    # 3.14 no longer auto-creates in the main thread (older Python did).
    try:
        asyncio.get_event_loop()
    except RuntimeError:
        asyncio.set_event_loop(asyncio.new_event_loop())

    judge_llm = LangchainLLMWrapper(
        ChatOllama(model=pipeline.config.generation.model_name, base_url=pipeline.config.generation.ollama_base_url)
    )
    judge_embeddings = LangchainEmbeddingsWrapper(
        HuggingFaceEmbeddings(model_name="BAAI/bge-small-en-v1.5")
    )
    run_config = RunConfig(max_workers=1)

    metrics = [context_precision, context_recall, faithfulness, answer_relevancy, answer_correctness]
    for metric in metrics:
        metric.llm = judge_llm
        if hasattr(metric, "embeddings"):
            metric.embeddings = judge_embeddings
        metric.init(run_config)

    detail_rows = []
    for i, example in enumerate(eval_set):
        print(f"  scoring row {i + 1}/{len(eval_set)}: {example['question']!r}")
        answer, contexts = pipeline.answer_with_contexts(example["question"])
        row = {
            "question": example["question"],
            "answer": answer,
            "contexts": contexts,
            "ground_truth": example["ground_truth_answer"],
        }
        detail_row = {
            "question": example["question"],
            "topic": example.get("topic"),
            "answer": answer,
            "ground_truth_answer": example["ground_truth_answer"],
            "contexts": "\n---\n".join(contexts),
        }
        for metric in metrics:
            score = metric.score(row)
            detail_row[metric.name] = score
            print(f"    {metric.name}: {score:.4f}")
        detail_rows.append(detail_row)

    mean_scores = {}
    for name in METRIC_NAMES:
        values = [row[name] for row in detail_rows if not math.isnan(row[name])]
        mean_scores[name] = sum(values) / len(values) if values else float("nan")

    return mean_scores, detail_rows


def safe_mean(values: list[float]) -> float:
    """Mean of the non-NaN values, or NaN if none are valid."""
    clean = [v for v in values if not math.isnan(v)]
    return sum(clean) / len(clean) if clean else float("nan")


def phase_summary(config: PipelineConfig) -> dict[str, object]:
    """Summarize which technique/setting each pipeline phase is using.

    Used for the header table written above each run's detail sheet and for
    the "Final Analysis" sheet's per-run columns, so it's easy to see at a
    glance what produced a given set of scores.

    Args:
        config: The PipelineConfig a run was executed with.

    Returns:
        An ordered mapping of phase name to the technique/setting used.
    """
    return {
        "chunking": config.chunking.default_strategy,
        "retrieval": config.retrieval.technique,
        "reranking": config.reranking.method,
        "mmr": config.diversification.enabled,
        "generation_backend": config.generation.backend,
        "generation_model": config.generation.model_name,
    }


def _composite_scores(mean_scores: dict[str, float]) -> dict[str, float]:
    correctness = safe_mean([mean_scores["faithfulness"], mean_scores["answer_correctness"]])
    completeness = mean_scores["context_recall"]
    return {
        "correctness": correctness,
        "completeness": completeness,
        "correctness_completeness": safe_mean([correctness, completeness]),
    }


def _unique_sheet_name(workbook, base_name: str) -> str:
    name = base_name[:31]
    if name not in workbook.sheetnames:
        return name
    for suffix in range(2, 1000):
        candidate = f"{base_name[:28]}_{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
    raise RuntimeError("Could not find a unique sheet name")


def append_run_to_workbook(
    workbook_path: Path,
    run_name: str,
    config: PipelineConfig,
    mean_scores: dict[str, float],
    detail_rows: list[dict],
) -> None:
    """Append one pipeline run's results to a shared Excel workbook.

    Each call adds a new sheet (named after ``run_name``) containing a phase
    summary table (which technique each phase used) followed by the
    per-question detail rows, and adds/updates a row for this run in the
    workbook's "Final Analysis" sheet (one row per run ever recorded, so you
    can compare runs over time without losing history).

    Args:
        workbook_path: Path to the .xlsx workbook (created if it doesn't exist).
        run_name: Short identifier for this run (used as the new sheet name,
            truncated/suffixed to stay unique and within Excel's 31-char limit).
        config: The PipelineConfig this run was executed with.
        mean_scores: Mean RAGAS metric scores, as returned by score_pipeline().
        detail_rows: Per-question detail rows, as returned by score_pipeline().
    """
    import openpyxl
    import pandas as pd
    from openpyxl.utils.dataframe import dataframe_to_rows

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = (
        openpyxl.load_workbook(workbook_path) if workbook_path.exists() else openpyxl.Workbook()
    )
    # A fresh Workbook() starts with one blank "Sheet"; drop it once we have
    # real content to write, so it doesn't linger as clutter.
    if workbook.sheetnames == ["Sheet"] and workbook["Sheet"].max_row == 1 and workbook["Sheet"].max_column == 1:
        del workbook["Sheet"]

    phases = phase_summary(config)
    composite = _composite_scores(mean_scores)

    # --- Per-run sheet: phase summary header table, then detail rows. ---
    sheet_name = _unique_sheet_name(workbook, run_name)
    sheet = workbook.create_sheet(sheet_name)

    sheet.append(["Phase", "Setting"])
    for phase, setting in phases.items():
        sheet.append([phase, setting])
    sheet.append([])
    sheet.append(["Metric", "Mean Score"])
    for metric_name in METRIC_NAMES:
        sheet.append([metric_name, mean_scores[metric_name]])
    for name, value in composite.items():
        sheet.append([name, value])
    sheet.append([])

    detail_df = pd.DataFrame(detail_rows)
    for row in dataframe_to_rows(detail_df, index=False, header=True):
        sheet.append(row)

    # --- Final Analysis sheet: one row per run, across all runs ever recorded. ---
    if "Final Analysis" not in workbook.sheetnames:
        analysis_sheet = workbook.create_sheet("Final Analysis")
        analysis_sheet.append(
            ["run_name", *phases.keys(), *METRIC_NAMES, *composite.keys()]
        )
    else:
        analysis_sheet = workbook["Final Analysis"]
    analysis_sheet.append(
        [sheet_name, *phases.values(), *(mean_scores[m] for m in METRIC_NAMES), *composite.values()]
    )

    workbook.save(workbook_path)
